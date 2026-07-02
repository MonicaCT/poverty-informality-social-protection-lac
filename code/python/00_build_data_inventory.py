"""Build a professional data inventory for the LAC poverty-informality project.

The script scans the local data archives named in the project brief, extracts
metadata from supported structured files, attaches nearby documentation and
scripts, and writes a Markdown inventory plus machine-readable companions.
"""

from __future__ import annotations

import csv
import json
import math
import os
import re
import sys
import unicodedata
from collections import Counter, defaultdict
from datetime import date
from pathlib import Path
from typing import Any

import pandas as pd

try:
    import openpyxl
except Exception:  # pragma: no cover - dependency is checked at runtime
    openpyxl = None

try:
    import pyarrow.feather as feather
    import pyarrow.parquet as pq
except Exception:  # pragma: no cover
    feather = None
    pq = None


ROOTS = [
    Path(r"D:\Papers Desarrollo_2026\Datos_papers_PART-I"),
    Path(r"C:\Users\Asus\Documents\Datos_papers_PART-II"),
]
PROJECT_DIR = Path(r"C:\Users\Asus\Documents\Github\poverty-informality-social-protection-lac")
METADATA_DIR = PROJECT_DIR / "data" / "metadata"
LOG_DIR = PROJECT_DIR / "logs"

DATA_EXTS = {
    ".csv",
    ".xlsx",
    ".xls",
    ".dta",
    ".sav",
    ".parquet",
    ".txt",
    ".rds",
    ".feather",
}
DOC_EXTS = {".pdf", ".doc", ".docx", ".md", ".json", ".yaml", ".yml"}
SCRIPT_EXTS = {".r", ".rmd", ".do", ".py", ".qmd", ".ipynb"}
ALL_EXTS = DATA_EXTS | DOC_EXTS | SCRIPT_EXTS

GENERATED_DATE = "2026-07-02"
CSV_EXACT_LINECOUNT_LIMIT_MB = 100
CSV_ESTIMATE_SAMPLE_BYTES = 8 * 1024 * 1024
MAX_SAMPLE_ROWS = 500
MAX_VARIABLE_PREVIEW = 60
MAX_RELATED_FILES = 8

COUNTRY_PATTERNS = {
    "ARG": ["argentina", "arg"],
    "BOL": ["bolivia", "bolivia (plurinational state of)", "bol", "bo"],
    "BRA": ["brazil", "brasil", "bra"],
    "CHL": ["chile", "chl"],
    "COL": ["colombia", "col"],
    "CRI": ["costa rica", "cri"],
    "CUB": ["cuba", "cub"],
    "DOM": ["dominican republic", "republica dominicana", "dom"],
    "ECU": ["ecuador", "ecu"],
    "SLV": ["el salvador", "salvador", "slv"],
    "GTM": ["guatemala", "gtm"],
    "HND": ["honduras", "hnd"],
    "HTI": ["haiti", "hti"],
    "JAM": ["jamaica", "jam"],
    "MEX": ["mexico", "mex"],
    "NIC": ["nicaragua", "nic"],
    "PAN": ["panama", "pan"],
    "PRY": ["paraguay", "pry"],
    "PER": ["peru", "per"],
    "URY": ["uruguay", "ury"],
    "VEN": ["venezuela", "ven"],
    "BLZ": ["belize", "blz"],
    "GUY": ["guyana", "guy"],
    "SUR": ["suriname", "sur"],
    "TTO": ["trinidad and tobago", "tto"],
}

ROLE_KEYWORDS = {
    "poverty": [
        "poverty",
        "pobre",
        "pobreza",
        "poor",
        "headcount",
        "extreme",
        "income",
        "ingreso",
        "consumption",
        "consumo",
        "welfare",
        "gini",
        "inequality",
        "equity",
        "sedlac",
    ],
    "labor_informality": [
        "informal",
        "informality",
        "labor",
        "labour",
        "empleo",
        "employment",
        "ocup",
        "unemployment",
        "desempleo",
        "ilostat",
        "work",
        "worker",
        "sector",
    ],
    "social_protection": [
        "social protection",
        "aspire",
        "transfer",
        "pension",
        "cash",
        "cct",
        "benefit",
        "bono",
        "subsid",
        "coverage",
        "program",
        "asistencia",
        "proteccion",
        "social",
    ],
    "gender": ["female", "male", "gender", "woman", "women", "mujer", "hombre", "sex"],
    "macro_controls": [
        "gdp",
        "pib",
        "inflation",
        "wdi",
        "world development",
        "government",
        "expenditure",
        "population",
        "urban",
        "health",
        "education",
        "school",
        "barro",
        "lee",
        "wgi",
    ],
    "demographics": [
        "age",
        "edad",
        "dependency",
        "population",
        "household",
        "hogar",
        "vivienda",
        "dhs",
        "census",
        "censo",
    ],
    "trade_finance": [
        "trade",
        "export",
        "import",
        "gravity",
        "financial",
        "finance",
        "commodity",
        "atlas",
        "tariff",
    ],
    "governance": ["governance", "institution", "wgi", "v-dem", "democracy", "regulation"],
    "climate_energy": ["energy", "emission", "climate", "faostat", "disaster", "fuel"],
}


def norm_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value)
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    return text.lower()


def md_escape(value: Any) -> str:
    text = "" if value is None else str(value)
    text = text.replace("\r", " ").replace("\n", " ")
    text = text.replace("|", "\\|")
    return text.strip()


def compact_list(values: list[Any] | tuple[Any, ...] | set[Any], limit: int = 12) -> str:
    vals = [str(v) for v in values if v is not None and str(v) != ""]
    if not vals:
        return ""
    if len(vals) <= limit:
        return "; ".join(vals)
    return "; ".join(vals[:limit]) + f"; ... (+{len(vals) - limit})"


def bytes_mb(size: int | float | None) -> float:
    if not size:
        return 0.0
    return round(float(size) / 1024 / 1024, 3)


def safe_stat(path: Path) -> dict[str, Any]:
    try:
        st = path.stat()
        return {"size": st.st_size, "size_mb": bytes_mb(st.st_size), "modified": date.fromtimestamp(st.st_mtime).isoformat()}
    except Exception:
        return {"size": 0, "size_mb": 0.0, "modified": ""}


def relative_to_root(path: Path) -> tuple[str, str, str, str]:
    for root in ROOTS:
        try:
            rel = path.relative_to(root)
            parts = rel.parts
            top = parts[0] if parts else path.name
            second = parts[1] if len(parts) > 1 else ""
            return str(root), str(rel), top, second
        except ValueError:
            continue
    return "", str(path), "", ""


def enumerate_files() -> list[Path]:
    files: list[Path] = []
    for root in ROOTS:
        if not root.exists():
            continue
        for path in root.rglob("*"):
            try:
                if path.is_file() and path.suffix.lower() in ALL_EXTS:
                    files.append(path)
            except OSError:
                continue
    return sorted(files, key=lambda p: str(p).lower())


def read_special_metadata() -> dict[str, dict[str, Any]]:
    special_path = METADATA_DIR / "r_inventory_special.json"
    if not special_path.exists():
        return {}
    try:
        data = json.loads(special_path.read_text(encoding="utf-8"))
    except Exception:
        return {}
    out: dict[str, dict[str, Any]] = {}
    for item in data:
        p = item.get("path", "")
        if p:
            out[Path(p).as_posix().lower()] = item
            out[str(Path(p)).lower()] = item
    return out


def detect_encoding(path: Path, sample_size: int = 131072) -> tuple[str, str]:
    raw = b""
    try:
        with path.open("rb") as f:
            raw = f.read(sample_size)
    except Exception as exc:
        return "utf-8", f"encoding_probe_failed: {exc}"
    for enc in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
        try:
            raw.decode(enc)
            return enc, ""
        except UnicodeDecodeError:
            continue
    return "latin-1", "encoding_fallback_latin1"


def sniff_delimiter(path: Path, encoding: str) -> str:
    try:
        with path.open("r", encoding=encoding, errors="replace", newline="") as f:
            sample = f.read(65536)
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
            return dialect.delimiter
        except Exception:
            first = sample.splitlines()[0] if sample.splitlines() else ""
            counts = {d: first.count(d) for d in [",", ";", "\t", "|"]}
            return max(counts, key=counts.get) if max(counts.values()) > 0 else ","
    except Exception:
        return ","


def count_or_estimate_rows(path: Path, size_mb: float) -> tuple[int | None, str]:
    try:
        if size_mb <= CSV_EXACT_LINECOUNT_LIMIT_MB:
            lines = 0
            with path.open("rb") as f:
                for chunk in iter(lambda: f.read(1024 * 1024), b""):
                    lines += chunk.count(b"\n")
            return max(lines - 1, 0), "exact_binary_line_count"
        with path.open("rb") as f:
            sample = f.read(CSV_ESTIMATE_SAMPLE_BYTES)
        line_count = sample.count(b"\n")
        if line_count <= 1:
            return None, "not_counted_large_file_insufficient_sample"
        avg_bytes = len(sample) / line_count
        estimate = max(int(path.stat().st_size / avg_bytes) - 1, 0)
        return estimate, f"estimated_from_first_{CSV_ESTIMATE_SAMPLE_BYTES // 1024 // 1024}MB"
    except Exception as exc:
        return None, f"row_count_failed: {exc}"


def read_delimited_sample(path: Path, size_mb: float) -> tuple[pd.DataFrame | None, dict[str, Any]]:
    encoding, enc_note = detect_encoding(path)
    delimiter = sniff_delimiter(path, encoding)
    meta = {"encoding": encoding, "delimiter": repr(delimiter), "notes": enc_note}
    try:
        df = pd.read_csv(
            path,
            nrows=MAX_SAMPLE_ROWS,
            sep=delimiter,
            encoding=encoding,
            on_bad_lines="skip",
            low_memory=False,
        )
        return df, meta
    except Exception as first_exc:
        try:
            df = pd.read_csv(
                path,
                nrows=MAX_SAMPLE_ROWS,
                sep=None,
                engine="python",
                encoding=encoding,
                on_bad_lines="skip",
            )
            meta["notes"] = (meta.get("notes", "") + f"; fallback_python_parser_after: {first_exc}").strip("; ")
            return df, meta
        except Exception as second_exc:
            meta["notes"] = f"delimited_sample_failed: {second_exc}"
            return None, meta


def inspect_delimited(path: Path, base: dict[str, Any]) -> tuple[dict[str, Any], list[dict[str, Any]]]:
    df, meta = read_delimited_sample(path, base["size_mb"])
    obs, method = count_or_estimate_rows(path, base["size_mb"])
    base.update(
        {
            "status": "ok" if df is not None else "metadata_only",
            "observations": obs,
            "observation_method": method,
            "encoding": meta.get("encoding", ""),
            "delimiter": meta.get("delimiter", ""),
            "notes": meta.get("notes", ""),
        }
    )
    if df is None:
        return base, []
    return finalize_from_sample(base, df)


def inspect_xlsx(path: Path, base: dict[str, Any]) -> tuple[dict[str, Any], list[dict[str, Any]]]:
    if openpyxl is None:
        base.update({"status": "metadata_only", "notes": "openpyxl_not_available"})
        return base, []
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        sheets = wb.sheetnames
        base["sheets"] = compact_list(sheets, 10)
        if not sheets:
            base.update({"status": "metadata_only", "notes": "workbook_has_no_sheets"})
            return base, []
        ws = wb[sheets[0]]
        max_row = ws.max_row or 0
        max_col = ws.max_column or 0
        rows = []
        for idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            rows.append(list(row))
            if idx >= min(max(max_row, 1), MAX_SAMPLE_ROWS + 5):
                break
        wb.close()
        header_idx = 0
        for i, row in enumerate(rows[:20]):
            non_empty = [cell for cell in row if cell is not None and str(cell).strip() != ""]
            if len(non_empty) >= 2:
                header_idx = i
                break
        header = rows[header_idx] if rows else []
        columns = []
        for idx, value in enumerate(header, start=1):
            name = str(value).strip() if value is not None and str(value).strip() != "" else f"unnamed_{idx}"
            columns.append(name)
        sample_rows = rows[header_idx + 1 : header_idx + 1 + MAX_SAMPLE_ROWS]
        df = pd.DataFrame(sample_rows, columns=columns[: len(sample_rows[0])] if sample_rows else columns)
        base.update(
            {
                "status": "ok",
                "observations": max(max_row - header_idx - 1, 0) if max_row else None,
                "observation_method": "openpyxl_sheet_dimension_first_sheet",
                "notes": f"first_sheet={sheets[0]}; workbook_sheets={len(sheets)}; max_col_first_sheet={max_col}",
            }
        )
        return finalize_from_sample(base, df)
    except Exception as exc:
        base.update({"status": "metadata_only", "notes": f"xlsx_inspection_failed: {exc}"})
        return base, []


def inspect_dta(path: Path, base: dict[str, Any]) -> tuple[dict[str, Any], list[dict[str, Any]]]:
    try:
        reader = pd.io.stata.StataReader(str(path), convert_categoricals=False)
        try:
            variables = list(reader.varlist)
            labels = reader.variable_labels() or {}
            obs = getattr(reader, "nobs", None)
            df = reader.read(nrows=MAX_SAMPLE_ROWS, convert_categoricals=False)
        finally:
            reader.close()
        base.update(
            {
                "status": "ok",
                "observations": int(obs) if obs is not None and not pd.isna(obs) else None,
                "observation_method": "stata_reader_metadata",
                "notes": "",
            }
        )
        result, var_rows = finalize_from_sample(base, df)
        for row in var_rows:
            row["variable_label"] = labels.get(row["variable_name"], "")
        if variables and not result.get("variables"):
            result["variables"] = compact_list(variables, MAX_VARIABLE_PREVIEW)
        return result, var_rows
    except Exception as exc:
        base.update({"status": "metadata_only", "notes": f"dta_inspection_failed: {exc}"})
        return base, []


def inspect_parquet(path: Path, base: dict[str, Any]) -> tuple[dict[str, Any], list[dict[str, Any]]]:
    if pq is None:
        base.update({"status": "metadata_only", "notes": "pyarrow_parquet_not_available"})
        return base, []
    try:
        pf = pq.ParquetFile(path)
        schema = pf.schema_arrow
        names = list(schema.names)
        obs = pf.metadata.num_rows if pf.metadata else None
        cols = names[: min(len(names), 80)]
        df = pf.read_row_group(0, columns=cols).to_pandas().head(MAX_SAMPLE_ROWS) if pf.num_row_groups else pd.DataFrame(columns=names)
        base.update(
            {
                "status": "ok",
                "observations": int(obs) if obs is not None else None,
                "observation_method": "parquet_file_metadata",
                "notes": f"row_groups={pf.num_row_groups}",
            }
        )
        result, var_rows = finalize_from_sample(base, df)
        result["variables"] = compact_list(names, MAX_VARIABLE_PREVIEW)
        result["n_variables"] = len(names)
        existing = {row["variable_name"] for row in var_rows}
        for field in schema:
            if field.name not in existing:
                var_rows.append(
                    {
                        "dataset_id": base["dataset_id"],
                        "variable_name": field.name,
                        "variable_type": str(field.type),
                        "variable_label": "",
                        "inferred_role": infer_variable_role(field.name),
                        "sample_missing_pct": "",
                    }
                )
        return result, var_rows
    except Exception as exc:
        base.update({"status": "metadata_only", "notes": f"parquet_inspection_failed: {exc}"})
        return base, []


def inspect_feather(path: Path, base: dict[str, Any]) -> tuple[dict[str, Any], list[dict[str, Any]]]:
    if feather is None:
        base.update({"status": "metadata_only", "notes": "pyarrow_feather_not_available"})
        return base, []
    try:
        table = feather.read_table(path)
        df = table.slice(0, MAX_SAMPLE_ROWS).to_pandas()
        base.update(
            {
                "status": "ok",
                "observations": table.num_rows,
                "observation_method": "feather_table_metadata",
                "notes": "",
            }
        )
        return finalize_from_sample(base, df)
    except Exception as exc:
        base.update({"status": "metadata_only", "notes": f"feather_inspection_failed: {exc}"})
        return base, []


def inspect_special(path: Path, base: dict[str, Any], special: dict[str, Any]) -> tuple[dict[str, Any], list[dict[str, Any]]]:
    variables = special.get("variables") or []
    labels = special.get("labels") or []
    types = special.get("variable_types") or []
    if isinstance(variables, str):
        variables = [variables]
    if isinstance(labels, str):
        labels = [labels]
    if isinstance(types, str):
        types = [types]
    base.update(
        {
            "status": special.get("status", "metadata_only"),
            "observations": special.get("observations"),
            "observation_method": special.get("observation_method") or special.get("observation_method".replace("_", "."), ""),
            "n_variables": special.get("n_variables") or len(variables),
            "variables": compact_list(variables, MAX_VARIABLE_PREVIEW),
            "sheets": compact_list(special.get("sheets") or [], 10),
            "notes": special.get("notes") or "",
        }
    )
    var_rows = []
    for idx, var in enumerate(variables):
        var_rows.append(
            {
                "dataset_id": base["dataset_id"],
                "variable_name": var,
                "variable_type": types[idx] if idx < len(types) else "",
                "variable_label": labels[idx] if idx < len(labels) else "",
                "inferred_role": infer_variable_role(var),
                "sample_missing_pct": "",
            }
        )
    augment_semantics(base, variable_names=list(variables), sample=None, labels=labels)
    return base, var_rows


def sample_missing_summary(df: pd.DataFrame) -> tuple[str, dict[str, Any]]:
    if df is None or df.empty:
        return "", {}
    try:
        missing = df.isna().mean() * 100
        overall = float(df.isna().sum().sum()) / max(float(df.shape[0] * df.shape[1]), 1.0) * 100
        high = [f"{col}:{missing[col]:.1f}%" for col in missing.index if missing[col] >= 50][:10]
        summary = f"sample_overall={overall:.1f}%"
        if high:
            summary += "; high_missing=" + "; ".join(high)
        return summary, {str(col): round(float(missing[col]), 2) for col in missing.index}
    except Exception:
        return "", {}


def infer_variable_role(name: str) -> str:
    text = norm_text(name)
    if re.search(r"(^|_)(iso3|iso|country|pais|economy|location|geo|region|depart|municip|province|state|admin)", text):
        return "spatial_identifier"
    if re.search(r"(^|_)(year|yr|anio|ano|date|fecha|time|period|wave|round|month|quarter)", text):
        return "time_identifier"
    if re.search(r"(id|folio|hhid|pid|person|household|hogar|vivienda|upm|psu|cluster)", text):
        return "survey_or_unit_identifier"
    for role, keywords in ROLE_KEYWORDS.items():
        if any(k in text for k in keywords):
            return role
    if re.search(r"(weight|pondera|factor|expansion|wgt)", text):
        return "survey_weight"
    return ""


def extract_years(variable_names: list[str], sample: pd.DataFrame | None, path_text: str) -> str:
    years: set[int] = set()
    for value in re.findall(r"(?:19|20)\d{2}", path_text):
        year = int(value)
        if 1900 <= year <= 2035:
            years.add(year)
    if sample is not None and not sample.empty:
        year_cols = [
            c
            for c in sample.columns
            if re.search(r"year|yr|anio|ano|fecha|date|time|period", norm_text(c))
        ]
        for col in year_cols[:8]:
            series = sample[col].dropna().astype(str).head(500)
            for item in series:
                for value in re.findall(r"(?:19|20)\d{2}", item):
                    year = int(value)
                    if 1900 <= year <= 2035:
                        years.add(year)
            numeric = pd.to_numeric(sample[col], errors="coerce")
            for value in numeric.dropna().head(500):
                if not math.isfinite(float(value)):
                    continue
                year = int(value)
                if 1900 <= year <= 2035:
                    years.add(year)
    if not years:
        return ""
    sorted_years = sorted(years)
    if len(sorted_years) == 1:
        return str(sorted_years[0])
    return f"{sorted_years[0]}-{sorted_years[-1]} ({len(sorted_years)} values)"


def extract_countries(variable_names: list[str], sample: pd.DataFrame | None, path_text: str) -> str:
    found: set[str] = set()
    text = norm_text(path_text + " " + " ".join(variable_names))
    for iso3, aliases in COUNTRY_PATTERNS.items():
        if any(re.search(rf"(^|[^a-z]){re.escape(alias)}([^a-z]|$)", text) for alias in aliases):
            found.add(iso3)
    if sample is not None and not sample.empty:
        ccols = [
            c
            for c in sample.columns
            if re.search(r"country|pais|iso3|iso|economy|location|nation|territory", norm_text(c))
        ]
        for col in ccols[:6]:
            values = [str(v).strip() for v in sample[col].dropna().unique().tolist()[:80]]
            for value in values:
                vnorm = norm_text(value)
                if re.fullmatch(r"[A-Z]{3}", value):
                    found.add(value)
                for iso3, aliases in COUNTRY_PATTERNS.items():
                    if any(alias in vnorm for alias in aliases):
                        found.add(iso3)
            if len(values) > 20 and not found:
                return f"multiple/global sample ({len(values)} unique values)"
    if not found:
        if re.search(r"latin america|lac|latam|america latina", text):
            return "LAC/multiple"
        if re.search(r"world|global|wdi|wid|ilostat|imf|un", text):
            return "global/multiple"
        return ""
    vals = sorted(found)
    if len(vals) > 12:
        return f"multiple ({len(vals)} ISO3 detected): " + "; ".join(vals[:12]) + "; ..."
    return "; ".join(vals)


def infer_identifiers(variable_names: list[str]) -> tuple[str, str, str]:
    spatial = []
    time = []
    unit = []
    for var in variable_names:
        text = norm_text(var)
        if re.search(r"iso3|iso|country|pais|economy|location|geo|region|depart|municip|province|state|admin|lat|lon|longitude|latitude|ubigeo|cod_", text):
            spatial.append(var)
        if re.search(r"year|yr|anio|ano|fecha|date|time|period|wave|round|month|quarter", text):
            time.append(var)
        if re.search(r"(^|_)(id|folio|hhid|pid|person|household|hogar|vivienda|upm|psu|cluster|sector|product|partner)", text):
            unit.append(var)
    return compact_list(spatial, 15), compact_list(time, 15), compact_list(unit, 15)


def infer_roles_and_uses(variable_names: list[str], labels: list[str] | None, path_text: str) -> tuple[list[str], str, str]:
    text = norm_text(path_text + " " + " ".join(variable_names) + " " + " ".join(labels or []))
    roles = []
    for role, keywords in ROLE_KEYWORDS.items():
        if any(k in text for k in keywords):
            roles.append(role)
    if not roles:
        roles = ["auxiliary_or_context"]
    uses = []
    if "poverty" in roles:
        uses.append("poverty and inequality indicators")
    if "labor_informality" in roles:
        uses.append("labor market and informality measures")
    if "social_protection" in roles:
        uses.append("social protection coverage, transfers, or programs")
    if "macro_controls" in roles:
        uses.append("macro and demographic controls")
    if "gender" in roles:
        uses.append("gender heterogeneity")
    if "demographics" in roles:
        uses.append("household, population, or survey aggregation")
    if "trade_finance" in roles:
        uses.append("external controls or robustness context")
    if "governance" in roles:
        uses.append("institutional controls")
    if "climate_energy" in roles:
        uses.append("contextual controls or vulnerability extensions")
    potential = "; ".join(uses) if uses else "documentation or auxiliary source pending manual review"
    return roles, "; ".join(roles), potential


def recommend_keys(spatial: str, time: str, unit: str, roles: list[str]) -> str:
    spatial_lower = norm_text(spatial)
    time_lower = norm_text(time)
    keys = []
    if ("iso3" in spatial_lower or "country" in spatial_lower or "pais" in spatial_lower) and (
        "year" in time_lower or "anio" in time_lower or "ano" in time_lower
    ):
        keys.append("iso3-country-year")
    elif spatial and time:
        keys.append("spatial_id-time_id")
    elif time:
        keys.append("time_id")
    if unit:
        keys.append("survey/unit identifiers for aggregation")
    if not keys and any(role in roles for role in ["macro_controls", "poverty", "social_protection", "labor_informality"]):
        keys.append("derive iso3/year from labels or source metadata")
    return "; ".join(keys)


def recommend_cleaning(extension: str, roles: list[str], spatial: str, time: str, notes: str) -> str:
    tasks = []
    if extension in {".csv", ".txt"}:
        tasks.append("validate delimiter, encoding, duplicate headers")
    if extension in {".xlsx", ".xls"}:
        tasks.append("standardize sheets and detect true header rows")
    if extension in {".dta", ".sav"}:
        tasks.append("preserve labels, missing-value codes, weights")
    if extension in {".parquet", ".feather"}:
        tasks.append("verify schema and row-group consistency")
    if "poverty" in roles:
        tasks.append("harmonize poverty definitions and currency/PPP basis")
    if "labor_informality" in roles:
        tasks.append("align employment/informality definitions")
    if "social_protection" in roles:
        tasks.append("separate coverage, incidence, benefits, and spending")
    if spatial:
        tasks.append("standardize ISO3 and country names")
    if time:
        tasks.append("standardize years/periods")
    if "estimated_from" in notes:
        tasks.append("run exact row count during production pipeline")
    return "; ".join(dict.fromkeys(tasks))


def integration_score(roles: list[str], countries: str, years: str, spatial: str, time: str, extension: str, path_text: str) -> int:
    score = 0
    if spatial:
        score += 2
    if time or years:
        score += 2
    if "poverty" in roles:
        score += 3
    if "labor_informality" in roles:
        score += 3
    if "social_protection" in roles:
        score += 3
    if "macro_controls" in roles:
        score += 2
    if "gender" in roles or "demographics" in roles:
        score += 1
    if "BOL" in countries or "bolivia" in norm_text(path_text):
        score += 2
    if any(token in norm_text(path_text) for token in ["sedlac", "aspire", "ilostat", "wdi", "gini", "ehogares", "panel_global"]):
        score += 3
    if extension in {".parquet", ".dta", ".rds"}:
        score += 1
    return score


def augment_semantics(
    base: dict[str, Any],
    variable_names: list[str],
    sample: pd.DataFrame | None,
    labels: list[str] | None = None,
) -> None:
    path_text = " ".join([base.get("relative_path", ""), base.get("dataset_name", ""), base.get("top_folder", "")])
    years = extract_years(variable_names, sample, path_text)
    countries = extract_countries(variable_names, sample, path_text)
    spatial, time, unit = infer_identifiers(variable_names)
    roles, role_text, potential = infer_roles_and_uses(variable_names, labels, path_text)
    keys = recommend_keys(spatial, time, unit, roles)
    cleaning = recommend_cleaning(base["extension"], roles, spatial, time, base.get("notes", ""))
    base.update(
        {
            "years": years,
            "countries": countries,
            "spatial_identifiers": spatial,
            "time_identifiers": time,
            "unit_identifiers": unit,
            "possible_merge_keys": keys,
            "possible_uses": role_text,
            "potential_research_uses": potential,
            "recommended_cleaning": cleaning,
            "integration_score": integration_score(roles, countries, years, spatial, time, base["extension"], path_text),
        }
    )


def finalize_from_sample(base: dict[str, Any], df: pd.DataFrame) -> tuple[dict[str, Any], list[dict[str, Any]]]:
    if df is None:
        df = pd.DataFrame()
    cols = [str(c) for c in df.columns]
    missing_summary, missing_by_col = sample_missing_summary(df)
    base["n_variables"] = len(cols)
    base["variables"] = compact_list(cols, MAX_VARIABLE_PREVIEW)
    base["missing_values"] = missing_summary
    augment_semantics(base, cols, df)
    var_rows = []
    for col in cols:
        try:
            dtype = str(df[col].dtype)
        except Exception:
            dtype = ""
        var_rows.append(
            {
                "dataset_id": base["dataset_id"],
                "variable_name": col,
                "variable_type": dtype,
                "variable_label": "",
                "inferred_role": infer_variable_role(col),
                "sample_missing_pct": missing_by_col.get(col, ""),
            }
        )
    return base, var_rows


def build_related_maps(files: list[Path]) -> tuple[dict[str, list[str]], dict[str, list[str]], dict[str, list[str]], dict[str, list[str]]]:
    docs_by_top: dict[str, list[str]] = defaultdict(list)
    scripts_by_top: dict[str, list[str]] = defaultdict(list)
    docs_by_dir: dict[str, list[str]] = defaultdict(list)
    scripts_by_dir: dict[str, list[str]] = defaultdict(list)
    for path in files:
        root, rel, top, _ = relative_to_root(path)
        if not root:
            continue
        ext = path.suffix.lower()
        rel_str = rel.replace("\\", "/")
        dir_key = str(path.parent).lower()
        if ext in DOC_EXTS:
            docs_by_top[top].append(rel_str)
            docs_by_dir[dir_key].append(rel_str)
        elif ext in SCRIPT_EXTS:
            scripts_by_top[top].append(rel_str)
            scripts_by_dir[dir_key].append(rel_str)
    return docs_by_top, scripts_by_top, docs_by_dir, scripts_by_dir


def base_record(path: Path, dataset_id: str) -> dict[str, Any]:
    stat = safe_stat(path)
    root, rel, top, second = relative_to_root(path)
    return {
        "dataset_id": dataset_id,
        "dataset_name": path.stem,
        "location": str(path),
        "root": root,
        "relative_path": rel.replace("\\", "/"),
        "top_folder": top,
        "subfolder": second,
        "extension": path.suffix.lower(),
        "size_mb": stat["size_mb"],
        "modified": stat["modified"],
        "status": "pending",
        "observations": None,
        "observation_method": "",
        "n_variables": None,
        "variables": "",
        "years": "",
        "countries": "",
        "missing_values": "",
        "possible_uses": "",
        "possible_merge_keys": "",
        "spatial_identifiers": "",
        "time_identifiers": "",
        "unit_identifiers": "",
        "recommended_cleaning": "",
        "potential_research_uses": "",
        "related_documentation": "",
        "related_scripts": "",
        "sheets": "",
        "encoding": "",
        "delimiter": "",
        "integration_score": 0,
        "notes": "",
    }


def inspect_data_file(path: Path, dataset_id: str, special_map: dict[str, dict[str, Any]]) -> tuple[dict[str, Any], list[dict[str, Any]]]:
    base = base_record(path, dataset_id)
    ext = path.suffix.lower()
    key1 = path.as_posix().lower()
    key2 = str(path).lower()
    if ext in {".xls", ".sav", ".rds"} and (key1 in special_map or key2 in special_map):
        return inspect_special(path, base, special_map.get(key1) or special_map.get(key2) or {})
    if ext in {".csv", ".txt"}:
        return inspect_delimited(path, base)
    if ext == ".xlsx":
        return inspect_xlsx(path, base)
    if ext == ".dta":
        return inspect_dta(path, base)
    if ext == ".parquet":
        return inspect_parquet(path, base)
    if ext == ".feather":
        return inspect_feather(path, base)
    base.update({"status": "metadata_only", "notes": "unsupported_or_unread_by_current_environment"})
    return base, []


def attach_related_files(
    row: dict[str, Any],
    docs_by_top: dict[str, list[str]],
    scripts_by_top: dict[str, list[str]],
    docs_by_dir: dict[str, list[str]],
    scripts_by_dir: dict[str, list[str]],
) -> None:
    path = Path(row["location"])
    dir_key = str(path.parent).lower()
    docs = docs_by_dir.get(dir_key) or docs_by_top.get(row["top_folder"], [])
    scripts = scripts_by_dir.get(dir_key) or scripts_by_top.get(row["top_folder"], [])
    row["related_documentation"] = compact_list(sorted(docs), MAX_RELATED_FILES)
    row["related_scripts"] = compact_list(sorted(scripts), MAX_RELATED_FILES)


def summarize_docs_scripts(files: list[Path]) -> list[dict[str, Any]]:
    rows = []
    for idx, path in enumerate(files, start=1):
        ext = path.suffix.lower()
        if ext not in DOC_EXTS and ext not in SCRIPT_EXTS:
            continue
        stat = safe_stat(path)
        root, rel, top, second = relative_to_root(path)
        rows.append(
            {
                "asset_id": f"A{idx:04d}",
                "asset_type": "documentation" if ext in DOC_EXTS else "script",
                "name": path.name,
                "location": str(path),
                "root": root,
                "relative_path": rel.replace("\\", "/"),
                "top_folder": top,
                "subfolder": second,
                "extension": ext,
                "size_mb": stat["size_mb"],
                "modified": stat["modified"],
                "possible_role": infer_doc_role(path),
            }
        )
    return rows


def infer_doc_role(path: Path) -> str:
    text = norm_text(" ".join(path.parts[-5:]))
    if any(k in text for k in ["codebook", "dictionary", "metadata", "diccionario", "manual", "questionnaire", "cuestionario"]):
        return "codebook_or_metadata"
    if path.suffix.lower() in SCRIPT_EXTS:
        return "existing_processing_script"
    if any(k in text for k in ["paper", "report", "method", "readme"]):
        return "methodological_documentation"
    return "documentation"


def write_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    if not rows:
        path.write_text("", encoding="utf-8")
        return
    fields = list(rows[0].keys())
    with path.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=fields, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def source_summary(rows: list[dict[str, Any]], assets: list[dict[str, Any]]) -> list[dict[str, Any]]:
    grouped: dict[tuple[str, str], dict[str, Any]] = {}
    for row in rows:
        key = (row["root"], row["top_folder"])
        item = grouped.setdefault(
            key,
            {
                "root": row["root"],
                "top_folder": row["top_folder"],
                "data_files": 0,
                "documentation_files": 0,
                "script_files": 0,
                "size_mb": 0.0,
                "extensions": Counter(),
                "roles": Counter(),
                "max_integration_score": 0,
                "recommended_use": "",
            },
        )
        item["data_files"] += 1
        item["size_mb"] += row.get("size_mb") or 0
        item["extensions"][row["extension"]] += 1
        for role in str(row.get("possible_uses", "")).split("; "):
            if role:
                item["roles"][role] += 1
        item["max_integration_score"] = max(item["max_integration_score"], int(row.get("integration_score") or 0))
    for asset in assets:
        key = (asset["root"], asset["top_folder"])
        item = grouped.setdefault(
            key,
            {
                "root": asset["root"],
                "top_folder": asset["top_folder"],
                "data_files": 0,
                "documentation_files": 0,
                "script_files": 0,
                "size_mb": 0.0,
                "extensions": Counter(),
                "roles": Counter(),
                "max_integration_score": 0,
                "recommended_use": "",
            },
        )
        if asset["asset_type"] == "documentation":
            item["documentation_files"] += 1
        else:
            item["script_files"] += 1
        item["extensions"][asset["extension"]] += 1
    out = []
    for item in grouped.values():
        roles = [role for role, _ in item["roles"].most_common(5)]
        recommended = recommend_source_use(item["top_folder"], roles, item["max_integration_score"])
        out.append(
            {
                "root": item["root"],
                "top_folder": item["top_folder"],
                "data_files": item["data_files"],
                "documentation_files": item["documentation_files"],
                "script_files": item["script_files"],
                "size_mb": round(item["size_mb"], 1),
                "extensions": "; ".join(f"{k}:{v}" for k, v in sorted(item["extensions"].items())),
                "main_roles": "; ".join(roles),
                "max_integration_score": item["max_integration_score"],
                "recommended_use": recommended,
            }
        )
    return sorted(out, key=lambda r: (r["root"], r["top_folder"]))


def recommend_source_use(top_folder: str, roles: list[str], score: int) -> str:
    text = norm_text(top_folder)
    if "aspire" in text:
        return "Core social protection source; harmonize program coverage and benefit/incidence variables."
    if "sedlac" in text or "equity" in text:
        return "Core poverty, inequality, and household welfare source for LAC indicators."
    if "ilostat" in text:
        return "Core labor market source; use for employment, unemployment, and informality indicators where available."
    if "wdi" in text or "gini" in text:
        return "Country-year controls and poverty/inequality benchmarks."
    if "ehogares" in text:
        return "Bolivia household survey microdata; aggregate to Bolivia-year and validate against regional series."
    if "panel" in text:
        return "Candidate integrated panel; audit provenance before using as analysis-ready backbone."
    if score >= 8:
        return "High-priority integration candidate for country-year panel."
    if score >= 5:
        return "Secondary controls, robustness, or extension source."
    return "Context, documentation, or lower-priority auxiliary source."


def obs_display(value: Any, method: str) -> str:
    if value is None or value == "" or (isinstance(value, float) and math.isnan(value)):
        return "NA"
    try:
        val = int(float(value))
        suffix = "~" if "estimated" in str(method) else ""
        return f"{suffix}{val:,}"
    except Exception:
        return str(value)


def make_markdown_table(rows: list[dict[str, Any]], fields: list[tuple[str, str]], limit: int | None = None) -> str:
    if limit is not None:
        rows = rows[:limit]
    header = "| " + " | ".join(label for _, label in fields) + " |"
    sep = "| " + " | ".join("---" for _ in fields) + " |"
    lines = [header, sep]
    for row in rows:
        vals = []
        for field, _ in fields:
            vals.append(md_escape(row.get(field, "")))
        lines.append("| " + " | ".join(vals) + " |")
    return "\n".join(lines)


def build_report(rows: list[dict[str, Any]], assets: list[dict[str, Any]], sources: list[dict[str, Any]]) -> str:
    ext_counts = Counter(row["extension"] for row in rows)
    role_counts = Counter()
    for row in rows:
        for role in str(row.get("possible_uses", "")).split("; "):
            if role:
                role_counts[role] += 1
    high_priority = sorted(rows, key=lambda r: (int(r.get("integration_score") or 0), r.get("size_mb") or 0), reverse=True)
    core = [
        row
        for row in high_priority
        if int(row.get("integration_score") or 0) >= 8
        or any(k in norm_text(row.get("relative_path", "")) for k in ["aspire", "sedlac", "ilostat", "ehogares", "panel_global", "wdi", "gini"])
    ]
    source_fields = [
        ("top_folder", "Source collection"),
        ("data_files", "Data files"),
        ("documentation_files", "Docs"),
        ("script_files", "Scripts"),
        ("size_mb", "Data MB"),
        ("extensions", "Formats"),
        ("main_roles", "Detected roles"),
        ("recommended_use", "Recommended use"),
    ]
    core_fields = [
        ("dataset_id", "ID"),
        ("dataset_name", "Dataset"),
        ("relative_path", "Location"),
        ("extension", "Format"),
        ("observations_display", "Obs."),
        ("n_variables", "Vars"),
        ("years", "Years"),
        ("countries", "Countries"),
        ("possible_uses", "Possible uses"),
        ("possible_merge_keys", "Merge keys"),
        ("recommended_cleaning", "Recommended cleaning"),
    ]
    catalog_fields = [
        ("dataset_id", "ID"),
        ("dataset_name", "Dataset"),
        ("relative_path", "Location"),
        ("extension", "Fmt"),
        ("observations_display", "Obs."),
        ("n_variables", "Vars"),
        ("variables", "Variables preview"),
        ("years", "Years"),
        ("countries", "Countries"),
        ("missing_values", "Missing values"),
        ("spatial_identifiers", "Spatial IDs"),
        ("time_identifiers", "Time IDs"),
        ("possible_merge_keys", "Merge keys"),
        ("potential_research_uses", "Research uses"),
        ("recommended_cleaning", "Cleaning"),
    ]
    for row in rows:
        row["observations_display"] = obs_display(row.get("observations"), row.get("observation_method", ""))
    method_notes = [
        "CSV/TXT observation counts are exact for files up to 100 MB and estimated from the first 8 MB for larger files.",
        "Parquet, Feather, Stata, RDS data frames, and `.xlsx` workbooks use format metadata where available.",
        "Older `.xls`, SPSS `.sav`, and `.rds` objects were enriched with R (`readxl`, `haven`, and `readRDS`).",
        "Missing-value diagnostics are sample-based unless the file metadata exposes full-file statistics.",
        "The companion variable catalog stores one row per detected variable; this Markdown report keeps variable previews readable.",
    ]
    lines = [
        "# Data Inventory",
        "",
        f"Generated: {GENERATED_DATE}",
        "",
        "Project: `poverty-informality-social-protection-lac`",
        "",
        "This inventory is the first project deliverable. It scans the two local archives recursively before any empirical modeling is selected or implemented.",
        "",
        "## Scope",
        "",
        "- Root 1: `D:\\Papers Desarrollo_2026\\Datos_papers_PART-I`",
        "- Root 2: `C:\\Users\\Asus\\Documents\\Datos_papers_PART-II`",
        f"- Structured data files inventoried: {len(rows):,}",
        f"- Documentation and existing scripts inventoried: {len(assets):,}",
        f"- Source collections detected: {len(sources):,}",
        "",
        "## File Census",
        "",
        make_markdown_table(
            [{"extension": ext, "count": count} for ext, count in sorted(ext_counts.items())],
            [("extension", "Format"), ("count", "Files")],
        ),
        "",
        "## Inventory Method",
        "",
    ]
    lines.extend(f"- {note}" for note in method_notes)
    lines.extend(
        [
            "",
            "## Source Collections",
            "",
            make_markdown_table(sources, source_fields),
            "",
            "## Core Integration Candidates",
            "",
            "These files and source collections are the highest-priority candidates for the research question because they contain country-year structure, poverty, informality/labor, social protection, macro controls, or Bolivia-specific survey information.",
            "",
            make_markdown_table(core[:120], core_fields),
            "",
            "## Research-Use Assessment",
            "",
            "- **Backbone panel candidate:** `panel_global_all_information.parquet` and related `PanelData_completo` files should be audited first because they may already consolidate many source-specific country-year indicators.",
            "- **Poverty and inequality:** prioritize SEDLAC, Equity Lab, WDI/GINI, WID, and household survey aggregates for monetary poverty, extreme poverty, and Gini indicators.",
            "- **Informality and labor:** prioritize ILOSTAT plus Bolivian household surveys for labor-market/informality constructs; validate definitions against survey codebooks and Stata/SPSS labels.",
            "- **Social protection:** prioritize World Bank ASPIRE and household-survey social transfer variables; separate coverage, adequacy, incidence, and expenditure concepts.",
            "- **Bolivia:** prioritize `52_EHogares_monica`, Bolivian ministry files, census files, DHS, and existing Bolivia aggregate parquet files for the Bolivia-vs-LAC comparison.",
            "- **Controls:** WDI, Barro-Lee, UNWPP, WHO, WGI, V-Dem, IMF, and fiscal/ministry datasets provide macro, demographic, education, health, governance, and public-finance controls.",
            "",
            "## Merge-Key Assessment",
            "",
            "The preferred analytical key is `iso3-year`. Files with subnational or household identifiers should be aggregated to country-year or modeled separately as Bolivia microdata validation layers. Product, partner, community, municipality, household, or person identifiers should never be merged directly into the cross-country panel without an explicit aggregation step.",
            "",
            "## Quality-Control Implications",
            "",
            "- Standardize ISO3 country codes and country names before any merge.",
            "- Normalize years and survey periods, especially files with waves, months, fiscal years, or date strings.",
            "- Preserve Stata/SPSS labels and survey missing-value codes before recoding.",
            "- Detect duplicate country-year observations after every source-specific aggregation.",
            "- Treat large estimated row counts as inventory approximations and run exact production counts in the formal pipeline.",
            "- Validate poverty, informality, and social protection definitions against source documentation before estimating models.",
            "",
            "## Machine-Readable Companions",
            "",
            "- `data/metadata/data_inventory_files.csv`: one row per structured data file with all required inventory fields.",
            "- `data/metadata/data_inventory_variables.csv`: one row per detected variable with type, label, role, and sample missingness.",
            "- `data/metadata/data_inventory_assets.csv`: documentation, codebooks, and scripts.",
            "- `data/metadata/data_inventory_sources.csv`: source-level summary and recommended use.",
            "- `data/metadata/data_inventory_files.json`: JSON version of the file inventory.",
            "",
            "## Complete Dataset Catalog",
            "",
            "The table below lists every structured data file detected recursively. Variable names are previews; use `data_inventory_variables.csv` for complete variable-level metadata.",
            "",
            make_markdown_table(rows, catalog_fields),
            "",
        ]
    )
    return "\n".join(lines)


def main() -> int:
    METADATA_DIR.mkdir(parents=True, exist_ok=True)
    LOG_DIR.mkdir(parents=True, exist_ok=True)
    files = enumerate_files()
    docs_by_top, scripts_by_top, docs_by_dir, scripts_by_dir = build_related_maps(files)
    special_map = read_special_metadata()
    data_files = [path for path in files if path.suffix.lower() in DATA_EXTS]
    rows: list[dict[str, Any]] = []
    variables: list[dict[str, Any]] = []
    failures: list[dict[str, Any]] = []
    for idx, path in enumerate(data_files, start=1):
        dataset_id = f"D{idx:05d}"
        try:
            row, var_rows = inspect_data_file(path, dataset_id, special_map)
            attach_related_files(row, docs_by_top, scripts_by_top, docs_by_dir, scripts_by_dir)
            rows.append(row)
            variables.extend(var_rows)
        except Exception as exc:
            row = base_record(path, dataset_id)
            row.update({"status": "error", "notes": str(exc)})
            attach_related_files(row, docs_by_top, scripts_by_top, docs_by_dir, scripts_by_dir)
            rows.append(row)
            failures.append({"dataset_id": dataset_id, "location": str(path), "error": repr(exc)})
        if idx % 100 == 0:
            print(f"Inventoried {idx}/{len(data_files)} files...", flush=True)
    assets = summarize_docs_scripts(files)
    sources = source_summary(rows, assets)
    write_csv(METADATA_DIR / "data_inventory_files.csv", rows)
    write_csv(METADATA_DIR / "data_inventory_variables.csv", variables)
    write_csv(METADATA_DIR / "data_inventory_assets.csv", assets)
    write_csv(METADATA_DIR / "data_inventory_sources.csv", sources)
    (METADATA_DIR / "data_inventory_files.json").write_text(
        json.dumps(rows, ensure_ascii=False, indent=2, default=str),
        encoding="utf-8",
    )
    failure_log = LOG_DIR / "data_inventory_failures.json"
    if failures:
        failure_log.write_text(
            json.dumps(failures, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
    elif failure_log.exists():
        failure_log.unlink()
    report = build_report(rows, assets, sources)
    (PROJECT_DIR / "DATA_INVENTORY.md").write_text(report, encoding="utf-8")
    print(f"Wrote {PROJECT_DIR / 'DATA_INVENTORY.md'}")
    print(f"Data files: {len(rows)}")
    print(f"Variables: {len(variables)}")
    print(f"Assets: {len(assets)}")
    print(f"Sources: {len(sources)}")
    print(f"Failures: {len(failures)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())



